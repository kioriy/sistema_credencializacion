"""
Adaptadores para archivos planos y Google Sheets.

- ``FileAdapter``: Lee CSV (auto-detect encoding) y Excel (.xlsx).
- ``GoogleSheetsAdapter``: Lee directamente de una hoja de Google Sheets
  usando gspread y credenciales de servicio.
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

from credencializacion.adapters.base import DataAdapter

logger = logging.getLogger(__name__)

# ── Codificaciones comunes en archivos mexicanos ─────────────────────
_ENCODINGS_TO_TRY = ("utf-8-sig", "utf-8", "latin-1", "cp1252")


# ═════════════════════════════════════════════════════════════════════
# FileAdapter — CSV y Excel
# ═════════════════════════════════════════════════════════════════════

class FileAdapter(DataAdapter):
    """Lee datos de archivos CSV o Excel (.xlsx).

    No instanciar directamente; usar el factory ``FileAdapter.from_file(path)``.

    Args:
        records: Lista de diccionarios ya parseados.
        columns: Lista de nombres de columna.
        source_path: Ruta al archivo fuente.
    """

    def __init__(
        self,
        records: list[dict[str, Any]],
        columns: list[str],
        source_path: Path,
    ) -> None:
        self._records = records
        self._columns = columns
        self._source_path = source_path

    # ── Interfaz DataAdapter ─────────────────────────────────────────

    def fetch_records(self, **kwargs) -> list[dict]:
        return list(self._records)

    def get_columns(self) -> list[str]:
        return list(self._columns)

    def get_source_name(self) -> str:
        return f"Archivo – {self._source_path.name}"

    # ── Factory ──────────────────────────────────────────────────────

    @classmethod
    def from_file(cls, path: Path) -> FileAdapter:
        """Detecta el formato y crea el adaptador adecuado.

        Args:
            path: Ruta al archivo CSV o Excel.

        Returns:
            Instancia de ``FileAdapter`` lista para usar.

        Raises:
            FileNotFoundError: Si el archivo no existe.
            ValueError: Si la extensión no es soportada.
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {path}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            return cls._from_csv(path)
        if suffix in (".xlsx", ".xls"):
            return cls._from_excel(path)
        raise ValueError(
            f"Formato de archivo no soportado: '{suffix}'. "
            f"Use .csv o .xlsx."
        )

    # ── CSV ──────────────────────────────────────────────────────────

    @classmethod
    def _from_csv(cls, path: Path) -> FileAdapter:
        """Lee un CSV con detección automática de codificación."""
        raw_bytes = path.read_bytes()
        content: str | None = None

        for encoding in _ENCODINGS_TO_TRY:
            try:
                content = raw_bytes.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if content is None:
            raise ValueError(
                f"No se pudo decodificar el archivo '{path.name}'. "
                f"Verifique la codificación."
            )

        reader = csv.DictReader(io.StringIO(content))
        columns = reader.fieldnames or []
        records = [dict(row) for row in reader]

        logger.info(
            "CSV '%s': %d registros, %d columnas (encoding detectado).",
            path.name, len(records), len(columns),
        )
        return cls(records=records, columns=list(columns), source_path=path)

    # ── Excel ────────────────────────────────────────────────────────

    @classmethod
    def _from_excel(cls, path: Path) -> FileAdapter:
        """Lee la primera hoja de un archivo Excel (.xlsx)."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "Se requiere 'openpyxl' para leer archivos Excel. "
                "Instale con: uv add openpyxl"
            ) from exc

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            wb.close()
            raise ValueError(f"El archivo '{path.name}' no tiene hojas activas.")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return cls(records=[], columns=[], source_path=path)

        # Primera fila = encabezados
        columns = [str(c).strip() if c is not None else f"col_{i}"
                   for i, c in enumerate(rows[0])]

        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {}
            for col_name, value in zip(columns, row):
                record[col_name] = value if value is not None else ""
            records.append(record)

        logger.info(
            "Excel '%s': %d registros, %d columnas.",
            path.name, len(records), len(columns),
        )
        return cls(records=records, columns=columns, source_path=path)


# ═════════════════════════════════════════════════════════════════════
# Helpers compartidos de autenticación (gspread)
# ═════════════════════════════════════════════════════════════════════

_SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


def load_service_account_credentials(credentials_path: Path):
    """Carga las credenciales de un service account de Google desde un JSON.

    Raises:
        FileNotFoundError: Si la ruta no existe.
        ValueError: Si el archivo no es un JSON de service account válido.
    """
    from google.oauth2.service_account import Credentials

    if not credentials_path.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo de credenciales: {credentials_path}"
        )
    try:
        return Credentials.from_service_account_file(
            str(credentials_path), scopes=_SHEETS_SCOPES,
        )
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            f"El archivo '{credentials_path.name}' no es un JSON de "
            f"service account de Google válido: {exc}"
        ) from exc


def get_service_account_email(credentials_path: Path) -> str:
    """Devuelve el ``client_email`` del service account (para mostrar al
    usuario qué correo debe tener acceso de lector al documento)."""
    creds = load_service_account_credentials(credentials_path)
    return getattr(creds, "service_account_email", "") or ""


# Nombre con el que se resguarda la credencial, para que la ruta guardada en
# los ajustes no dependa del nombre aleatorio que Google le pone al descargarla.
CREDENCIAL_SHEETS_NOMBRE = "google-sheets.json"


def instalar_credencial_sheets(origen: Path) -> Path:
    """Resguarda el JSON de service account en la carpeta de la app.

    La app guarda en sus ajustes la RUTA del archivo, no una copia, así que
    dejarlo en Descargas la rompe en cuanto el sistema limpia esa carpeta.
    Esta función lo lleva a una ubicación estable propia del sistema operativo
    (``get_credentials_dir()``), con permisos de solo-el-dueño.

    El archivo se **valida antes de moverse**: si no es un JSON de service
    account, no se toca nada. Si ya había una credencial instalada, se conserva
    como ``google-sheets.json.anterior`` en vez de perderse.

    Args:
        origen: Archivo elegido por el usuario.

    Returns:
        Ruta definitiva del archivo resguardado.

    Raises:
        FileNotFoundError: Si el archivo de origen no existe.
        ValueError: Si no es un JSON de service account válido.
        OSError: Si no se pudo escribir en la carpeta de credenciales.
    """
    import shutil

    from credencializacion.utils.paths import get_credentials_dir

    origen = Path(origen).expanduser()
    # Valida antes de mover: si el archivo no sirve, el usuario se queda con
    # su descarga intacta y con un mensaje que dice por qué.
    load_service_account_credentials(origen)

    destino = get_credentials_dir() / CREDENCIAL_SHEETS_NOMBRE
    if origen.resolve() == destino.resolve():
        return destino  # ya está resguardado; nada que hacer

    if destino.exists():
        anterior = destino.with_suffix(destino.suffix + ".anterior")
        try:
            if anterior.exists():
                anterior.unlink()
            destino.replace(anterior)
        except OSError as exc:  # noqa: BLE001
            logger.warning("No se pudo conservar la credencial previa: %s", exc)

    shutil.copy2(origen, destino)
    try:
        destino.chmod(0o600)
    except OSError:
        pass  # Windows y algunos sistemas de archivos no soportan chmod

    # El original se retira solo después de que la copia quedó en su sitio.
    # Si no se puede (permisos, archivo en uso), no es motivo de error: lo que
    # importa es que la app ya tiene su copia estable.
    try:
        origen.unlink()
    except OSError as exc:  # noqa: BLE001
        logger.info(
            "La credencial se resguardó en %s; el original en %s no se pudo "
            "retirar: %s", destino, origen, exc,
        )

    logger.info("Credencial de Google Sheets resguardada en %s", destino)
    return destino


def authorize_gspread_client(credentials_path: Path):
    """Autoriza un cliente gspread con credenciales de service account."""
    try:
        import gspread
    except ImportError as exc:
        raise ImportError(
            "Se requieren 'gspread' y 'google-auth' para acceder a "
            "Google Sheets. Instale con: uv add gspread google-auth"
        ) from exc

    creds = load_service_account_credentials(credentials_path)
    return gspread.authorize(creds)


def open_spreadsheet_by_name(client, document_name: str):
    """Abre (por título) el documento de Google Sheets indicado.

    Raises:
        ConnectionError: Si no se encuentra un documento con ese nombre
                         visible para el service account.
    """
    import gspread

    try:
        return client.open(document_name)
    except gspread.SpreadsheetNotFound:
        raise ConnectionError(
            f"No se encontró un documento de Google Sheets llamado "
            f"'{document_name}'. Verifique el nombre y que el documento "
            f"haya sido compartido con el correo del service account."
        )


# ═════════════════════════════════════════════════════════════════════
# GoogleSheetsAdapter — gspread
# ═════════════════════════════════════════════════════════════════════

class GoogleSheetsAdapter(DataAdapter):
    """Lee datos directamente de una hoja de Google Sheets.

    Requiere un archivo de credenciales de servicio (service account JSON).

    Args:
        spreadsheet_id: ID de la hoja de cálculo de Google.
        sheet_name: Nombre de la pestaña (por defecto, la primera).
        credentials_path: Ruta al JSON de credenciales del service account.
    """

    def __init__(
        self,
        spreadsheet_id: str,
        sheet_name: str | None = None,
        credentials_path: Path | None = None,
    ) -> None:
        self._spreadsheet_id = spreadsheet_id
        self._sheet_name = sheet_name
        self._credentials_path = credentials_path
        self._records: list[dict[str, Any]] = []
        self._columns: list[str] = []
        self._fetched = False

    # ── Interfaz DataAdapter ─────────────────────────────────────────

    def fetch_records(self, **kwargs) -> list[dict]:
        """Descarga los registros de Google Sheets.

        Keyword Args:
            force_refresh: Si True, re-descarga aunque ya se hayan
                           obtenido previamente.
        """
        force = kwargs.get("force_refresh", False)
        if not self._fetched or force:
            self._load_from_sheets()
        return list(self._records)

    def get_columns(self) -> list[str]:
        if not self._fetched:
            self._load_from_sheets()
        return list(self._columns)

    def get_source_name(self) -> str:
        label = self._sheet_name or "Hoja 1"
        return f"Google Sheets – {label}"

    # ── Lógica interna ───────────────────────────────────────────────

    def _load_from_sheets(self) -> None:
        """Carga datos usando gspread con credenciales de servicio."""
        try:
            import gspread
            from google.oauth2.service_account import Credentials
        except ImportError as exc:
            raise ImportError(
                "Se requieren 'gspread' y 'google-auth' para acceder a "
                "Google Sheets. Instale con: uv add gspread google-auth"
            ) from exc

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]

        if self._credentials_path and self._credentials_path.exists():
            creds = Credentials.from_service_account_file(
                str(self._credentials_path), scopes=scopes,
            )
        else:
            # Intentar credenciales por defecto del entorno
            try:
                from google.auth import default as google_default
                creds, _ = google_default(scopes=scopes)
            except Exception as exc:
                raise ConnectionError(
                    "No se encontraron credenciales de Google. "
                    "Proporcione un archivo de credenciales de servicio."
                ) from exc

        client = gspread.authorize(creds)

        try:
            spreadsheet = client.open_by_key(self._spreadsheet_id)
        except gspread.SpreadsheetNotFound:
            raise ConnectionError(
                f"No se encontró la hoja de cálculo con ID: "
                f"'{self._spreadsheet_id}'. Verifique que el service "
                f"account tenga acceso."
            )

        if self._sheet_name:
            try:
                worksheet = spreadsheet.worksheet(self._sheet_name)
            except gspread.WorksheetNotFound:
                raise ValueError(
                    f"No se encontró la pestaña '{self._sheet_name}' en la "
                    f"hoja de cálculo."
                )
        else:
            worksheet = spreadsheet.sheet1

        all_values = worksheet.get_all_records(
            expected_headers=None,
            default_blank="",
        )

        if all_values:
            self._columns = list(all_values[0].keys())
            # Asegurar que todos los valores sean strings para consistencia
            self._records = [
                {k: str(v) if v is not None else "" for k, v in row.items()}
                for row in all_values
            ]
        else:
            self._columns = []
            self._records = []

        self._fetched = True

        logger.info(
            "Google Sheets '%s': %d registros, %d columnas.",
            self._sheet_name or "Hoja 1",
            len(self._records),
            len(self._columns),
        )
